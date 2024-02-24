# 필요한 모듈 임포트
import pymysql
import openpyxl
import requests
# 마리아DB와 연결
conn = pymysql.connect(host='siriens.mycafe24.com', user='siriens', password='hosting1004!', db='siriens', charset='utf8')
# 커서 객체 생성
cursor = conn.cursor()
# 쿼리문 실행
sql = """SELECT '', yyyy, mm, dd, day, signal_grp, stock_cd, stock_nm, '', close_fg, close_rate, volume, trading_value
        , open, high, low, title, link
        FROM wave_evening
        ORDER BY yyyy, mm, dd, signal_grp, close_rate DESC"""
cursor.execute(sql)
# 결과를 모두 가져오기
result = cursor.fetchall()
# 커서와 연결 닫기
cursor.close()
conn.close()

# 엑셀 파일 생성
wb = openpyxl.Workbook()
# 첫 번째 시트 활성화
ws = wb.active
# 결과를 엑셀 파일에 쓰기
for row in result:
    ws.append(row)
# title 컬럼에 하이퍼링크 걸기
for i in range(2, ws.max_row + 1):
    # close_rate 컬럼의 셀 서식을 소수점 둘째자리까지 표시하도록 변경
    ws.cell(row=i, column=10).number_format = "0.00"
    # amount 컬럼의 셀 서식을 정수로 표시하도록 변경
    ws.cell(row=i, column=13).number_format = "0"
    ws.cell(row=i, column=17).hyperlink = ws.cell(row=i, column=18).value
    ws.cell(row=i, column=17).style = "Hyperlink"
# 엑셀 파일 저장
wb.save("E:/Project/202410/data/월웨엑셀작업/(sophia)wave_evening.xlsx")
