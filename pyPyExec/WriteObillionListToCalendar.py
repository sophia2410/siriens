import pymysql
import openpyxl
import calendar
import configparser

# 설정 파일 읽기
config = configparser.ConfigParser()
config.read('E:/Project/202410/www/boot/common/db/database_config.ini')

# MariaDB 연결
db = pymysql.connect(
    host=config.get('database', 'host'),
    user=config.get('database', 'user'),
    password=config.get('database', 'password'),
    db=config.get('database', 'db'),
    charset=config.get('database', 'charset')
)

# 커서 생성
cursor = db.cursor()

# 달력을 만들기 위한 연도와 월을 설정합니다.
year = 2023
month = 9

# SQL 쿼리를 작성합니다.
sql = f"""
SELECT '' ,'', ob.name, '', '', '', '', '', ob.code, ob.date, ob.amount
from (SELECT b.code, b.name, a.date, round(a.amount/100000000,0) amount, a.close_rate FROM daily_price a
inner join stock b
on b.code = a.code
and b.last_yn = 'Y'
WHERE a.date BETWEEN '{year}{month:02d}01' AND '{year}{month:02d}31' -- 날짜 범위를 지정합니다.
AND amount >= 100000000000 -- amount가 1천억 이상인 조건을 추가합니다.
AND high >= pre_close * 1.15 -- high가 pre_close의 15% 이상 상승한 조건을 추가합니다.
AND close >= open * 1.09 -- close가 open의 9% 이상 상승한 조건을 추가합니다.
) ob
ORDER BY date, amount DESC
"""

# SQL 쿼리를 실행합니다.
cursor.execute(sql)

# SQL 쿼리의 결과를 가져옵니다.
result = cursor.fetchall()

# 데이터베이스 연결을 닫습니다.
db.close()

# 엑셀 파일을 생성합니다.
wb = openpyxl.Workbook()

# 엑셀 파일의 첫 번째 시트를 활성화합니다.
ws = wb.active

# 엑셀 파일의 시트 이름을 변경합니다.
ws.title = "종목 정보"

# 달력 객체를 생성합니다.
cal = calendar.Calendar()

# 달력을 만들기 위한 행과 열의 개수를 설정합니다.
rows = 13
cols = 5

# 달력을 만들기 위한 시작 행과 열을 설정합니다.
start_row = 1
start_col = 1

# 달력의 제목을 작성합니다.
ws.cell(row=start_row, column=start_col, value=f"{year}년 {month}월")

# A열의 너비를 20으로 설정합니다.
ws.column_dimensions['A'].width = 39
ws.column_dimensions['B'].width = 39
ws.column_dimensions['C'].width = 39
ws.column_dimensions['D'].width = 39
ws.column_dimensions['E'].width = 39


# 월화수목금까지만 표시합니다.
days = ["월", "화", "수", "목", "금"]
for i in range(len(days)):
    ws.cell(row=start_row + 1, column=start_col + i, value=days[i])

# 달력의 날짜를 작성합니다.
for i, day in enumerate(cal.itermonthdays(year, month)):
    # 0이 아닌 경우에만 셀에 값을 넣습니다.
    if day != 0:
        ws.cell(row=start_row + 2 + (i // cols) * 2, column=start_col + i % cols, value=day)

# 달력의 날짜에 해당하는 종목 정보를 작성합니다.
for row in result:
    # 종목 정보의 날짜를 파싱합니다.
    date = str(row[9].decode('utf-8'))
    day = int(date[6:])

    # 종목 정보의 종목명과 종목코드를 가져옵니다.
    name = row[2].decode('utf-8')
    amount = str(row[10]) + '억'

    # 달력의 날짜와 일치하는 셀을 찾습니다.
    for r in range(start_row + 2, start_row + 2 + rows):
        for c in range(start_col, start_col + cols):
            cell = ws.cell(row=r, column=c)
            if cell.value == day:
                # 셀의 아래 셀에 종목 정보를 추가합니다.# 줄바꿈 함수를 사용하여 셀 내용을 줄바꿈합니다.
                if ws.cell(row=r + 1, column=c).value is not None:
                    ws.cell(row=r + 1, column=c, value=f"{ws.cell(row=r + 1, column=c).value}\n{name}({amount})")
                else:
                    ws.cell(row=r + 1, column=c, value=f"{name}({amount})")
                    
                # 1행의 높이를 40으로 설정합니다.
                ws.row_dimensions[r + 1].height = 200
                break
            
# 엑셀 파일을 저장합니다.
month = str(month).zfill(2)
wb.save(f"E:/★2030100★ 꿈은 이루어진다/★ 유목민 ★/W 오빌리언/stock_calendar_{year}{month}.xlsx")

# 엑셀 파일을 닫습니다.
wb.close()

# 코드가 완료되었음을 알립니다.
print("코드가 완료되었습니다.")