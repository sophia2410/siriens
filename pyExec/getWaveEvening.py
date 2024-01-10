# 필요한 모듈 임포트
import pymysql
import openpyxl
import requests
# 마리아DB와 연결
conn = pymysql.connect(host='yunseul0907.cafe24.com', user='yunseul0907', password='hosting1004!', db='yunseul0907', charset='utf8')
# 커서 객체 생성
cursor = conn.cursor()
# 쿼리문 실행
sql = """SELECT '', B.yyyy, B.mm, B.dd
, CASE B.day
    WHEN 1 THEN '일'
    WHEN 2 THEN '월'
    WHEN 3 THEN '화'
    WHEN 4 THEN '수'
    WHEN 5 THEN '목'
    WHEN 6 THEN '금'
    WHEN 7 THEN '토'
  END AS day
, A.signal_grp, A.stock_last, ''
, CASE WHEN C.close_rate > 29.5 THEN '↑' ELSE CASE WHEN  C.close_rate > 0 THEN '▲' ELSE CASE WHEN C.close_rate < -29.5 THEN '↓' ELSE '▼' END END END AS plus, C.close_rate, C.volume, round(C.amount/1000000,0) amount
, C.open, C.high, C.low
, A.title, A.link
FROM wave_evening_edit3 A
INNER JOIN calendar B
ON B.date = A.page_date
INNER JOIN daily_price C
ON C.date = A.page_date
AND C.code = A.code_last
WHERE A.title != ''
ORDER BY B.yyyy, B.mm, B.dd, B.day, A.signal_grp, C.close_rate DESC"""
cursor.execute(sql)
# 결과를 모두 가져오기
result = cursor.fetchall()
# 커서와 연결 닫기
cursor.close()
conn.close()

# 엑셀 파일 생성
wb = openpyxl.Workbook()
# 첫 번째 시트 활성화
ws = wb.active
# 결과를 엑셀 파일에 쓰기
for row in result:
    ws.append(row)
# title 컬럼에 하이퍼링크 걸기
for i in range(2, ws.max_row + 1):
    # close_rate 컬럼의 셀 서식을 소수점 둘째자리까지 표시하도록 변경
    ws.cell(row=i, column=10).number_format = "0.00"
    # amount 컬럼의 셀 서식을 정수로 표시하도록 변경
    ws.cell(row=i, column=12).number_format = "0"
    ws.cell(row=i, column=16).hyperlink = ws.cell(row=i, column=17).value
    ws.cell(row=i, column=16).style = "Hyperlink"
# 엑셀 파일 저장
wb.save("result.xlsx")
